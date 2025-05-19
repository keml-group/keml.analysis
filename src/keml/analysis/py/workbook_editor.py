from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import numpy as np

def fTi_fTa_columns_colors(ws):
    red_fill = PatternFill(start_color='FF5F5F', end_color='FF5F5F', fill_type='solid')
    green_fill = PatternFill(start_color='339966', end_color='339966', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF59E', end_color='FFF59E', fill_type='solid')
    ws.conditional_formatting.add(f'E3:F{len(ws["A"])}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add(f'E3:F{len(ws["A"])}', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add(f'E3:F{len(ws["A"])}', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add(f'E3:F{len(ws["A"])}', FormulaRule(formula=[f'NOT(ISNUMBER(O3:E{len(ws["A"])}))'], stopIfTrue=True, fill=yellow_fill))

def insert_rand_values(path):
    wb = load_workbook(path)
    ws = wb['Trust']
    fTi_fTa_columns_colors(ws)
    rand_i = 2 * np.random.rand(len(ws['A'])) - 1
    rand_a = 2 * np.random.rand(len(ws['A'])) - 1
    for i in range(3, len(ws['A']) + 1):
        ws[f'E{i}'].value = rand_i[i - 3]
        ws[f'F{i}'].value = rand_a[i - 3]
    wb.save(path)
    wb.close()

def insert_fTi(value, info_id, path):
    wb = load_workbook(path)
    ws = wb['Trust']
    fTi_fTa_columns_colors(ws)
    if not -1 <= value <= 1:
        wb.save(path)
        wb.close()
        raise ValueError(f'The given value must be a number between -1 and 1')
    if info_id + 2 > len(ws['A']):
        wb.save(path)
        wb.close()
        raise IndexError(f'The given info ID {info_id} is out of range')
    ws[f'E{info_id + 2}'].value = value
    wb.save(path)
    wb.close()

def insert_fTa(value, info_id, path):
    wb = load_workbook(path)
    ws = wb['Trust']
    fTi_fTa_columns_colors(ws)
    if not -1 <= value <= 1:
        wb.save(path)
        wb.close()
        raise ValueError(f'The given value must be a number between -1 and 1')
    if info_id + 2 > len(ws['A']):
        wb.save(path)
        wb.close()
        raise IndexError(f'The given info ID {info_id} is out of range')
    ws[f'F{info_id + 2}'].value = value
    wb.save(path)
    wb.close()