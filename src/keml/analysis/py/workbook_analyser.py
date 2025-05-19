from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import os

def add_new_columns(ws):
    ws.merge_cells('O1:P1')
    ws.merge_cells('Q1:R1')
    ws.merge_cells('S1:T1')
    ws.merge_cells('U1:V1')
    ws['O1'].font = Font(size=18)
    ws['O1'] = "diffs(a)"
    ws['O2'] = "fTi - T(a)"
    ws['P2'] = "fTa - T(a)"
    ws['Q1'].font = Font(size=18)
    ws['Q1'] = "diffs(b)"
    ws['Q2'] = "fTi - T(b)"
    ws['R2'] = "fTa - T(b)"
    ws['S1'].font = Font(size=18)
    ws['S1'] = "diffs(c)"
    ws['S2'] = "fTi - T(c)"
    ws['T2'] = "fTa - T(c)"
    ws['U1'].font = Font(size=18)
    ws['U1'] = "diffs(d)"
    ws['U2'] = "fTi - T(d)"
    ws['V2'] = "fTa - T(d)"
    b1 = Side(border_style="medium", color="000000")
    b2 = Side(border_style="medium", color="FFFFFF")
    red_fill = PatternFill(start_color='FF5F5F', end_color='FF5F5F', fill_type='solid')
    green_fill = PatternFill(start_color='339966', end_color='339966', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF59E', end_color='FFF59E', fill_type='solid')
    ws.conditional_formatting.add(f'O3:V{len(ws["A"])}', CellIsRule(operator='notBetween', formula=['-0.5', '0.5'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add(f'O3:V{len(ws["A"])}', CellIsRule(operator='between', formula=['-0.5', '0.5'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add(f'O3:V{len(ws["A"])}', FormulaRule(formula=[f'NOT(ISNUMBER(O3:V{len(ws["A"])}))'], stopIfTrue=True, fill=yellow_fill))
    for cells in ws['O1:V2']:
        for i in range(len(cells)):
            if (i % 2 == 0):
                cells[i].border = Border(left=b1)
            cells[i].alignment = Alignment(horizontal='center')
    for cells in ws[f'O3:V{len(ws["A"])}']:
        for i in range(len(cells)):
            if (i % 2 == 0):
                cells[i].border = Border(left=b2)
            cells[i].alignment = Alignment(horizontal='center')
            cells[i].number_format = "+0.00;-0.00;0.00"

def fill_diff_columns(ws):
    j = 3
    for cells in ws[f'O3:V{len(ws["A"])}']:
        if ws[f'E{j}'].value != '#NUM!':
            cells[0].value = ws[f'E{j}'].value - ws[f'H{j}'].value         
            cells[2].value = ws[f'E{j}'].value - ws[f'J{j}'].value        
            cells[4].value = ws[f'E{j}'].value - ws[f'L{j}'].value           
            cells[6].value = ws[f'E{j}'].value - ws[f'N{j}'].value           
        else:
            for i in range(0, len(cells), 2):
                cells[i].value = '#NUM!'
        if ws[f'F{j}'].value != '#NUM!':
            cells[1].value = ws[f'F{j}'].value - ws[f'H{j}'].value
            cells[3].value = ws[f'F{j}'].value - ws[f'J{j}'].value
            cells[5].value = ws[f'F{j}'].value - ws[f'L{j}'].value
            cells[7].value = ws[f'F{j}'].value - ws[f'N{j}'].value
        else:
            for i in range(1, len(cells), 2):
                cells[i].value = '#NUM!'
        j += 1

def run_wb_analysis(path_f, path_h):
    wb = load_workbook(path_f)
    ws = wb['Trust']
    add_new_columns(ws)
    fill_diff_columns(ws)
    create_stats(ws, path_h) 
    wb.save(path_f)
    wb.close()

def create_histograms(path_t_h, diffs_fT, key):
    fig = plt.figure(figsize=(10,4))
    axs = fig.subplots(1, 2, sharey=True)
    axs[0].xaxis.set_major_locator(ticker.MultipleLocator(0.5))
    axs[1].xaxis.set_major_locator(ticker.MultipleLocator(0.5))
    axs[0].yaxis.get_major_locator().set_params(integer=True)
    axs[1].yaxis.get_major_locator().set_params(integer=True)
    axs[0].set_title(f"fTi - T({key})")
    axs[1].set_title(f"fTa - T({key})")
    _, _, patches0 = axs[0].hist(diffs_fT[key][0], bins=8, range=(-2,2))
    _, _, patches1 = axs[1].hist(diffs_fT[key][1], bins=8, range=(-2,2)) 
    i = 0
    for patch0, patch1 in zip(patches0, patches1):
        if i == 0 or i == 7:
            patch0.set_facecolor('red')
            patch1.set_facecolor('red')
        elif i == 1 or i == 6:
            patch0.set_facecolor('orange')
            patch1.set_facecolor('orange')
        elif i == 2 or i == 5:
            patch0.set_facecolor('yellow')
            patch1.set_facecolor('yellow')
        elif i == 3 or i == 4:
            patch0.set_facecolor('green')
            patch1.set_facecolor('green')
        i += 1
    key_path = os.path.join(path_t_h, f"hist_{key}.jpg")
    plt.savefig(key_path)
    plt.close()

def create_stats(ws, path_t_h):
    diffs_fT = {'a':[None,None], 'b':[None,None], 'c':[None,None], 'd':[None,None]}
    abs_diffs_fT = {'a':[None,None], 'b':[None,None], 'c':[None,None], 'd':[None,None]}
    mean_abs_diffs_fT = {'a':[None,None], 'b':[None,None], 'c':[None,None], 'd':[None,None]}
    var_abs_diffs_fT = {'a':[None,None], 'b':[None,None], 'c':[None,None], 'd':[None,None]}
    std_abs_diffs_fT = {'a':[None,None], 'b':[None,None], 'c':[None,None], 'd':[None,None]}
    columns = {'a':['O', 'P'], 'b':['Q', 'R'], 'c':['S', 'T'], 'd':['U', 'V']}   
    l = len(ws["A"])
    for key in diffs_fT:
        diffs_fT[key][0] = [cells[0].value for cells in ws[f'{columns[key][0]}3:{columns[key][0]}{l}']]
        diffs_fT[key][1] = [cells[0].value for cells in ws[f'{columns[key][1]}3:{columns[key][1]}{l}']]
        if not os.path.exists(path_t_h):
            os.mkdir(path_t_h)
        create_histograms(path_t_h, diffs_fT, key)
        abs_diffs_fT[key][0] = np.abs(diffs_fT[key][0])
        abs_diffs_fT[key][1] = np.abs(diffs_fT[key][1])
        mean_abs_diffs_fT[key][0] = np.mean(abs_diffs_fT[key][0])
        mean_abs_diffs_fT[key][1] = np.mean(abs_diffs_fT[key][1])
        var_abs_diffs_fT[key][0] = np.var(abs_diffs_fT[key][0])
        var_abs_diffs_fT[key][1] = np.var(abs_diffs_fT[key][1])
        std_abs_diffs_fT[key][0] = np.std(abs_diffs_fT[key][0])
        std_abs_diffs_fT[key][1] = np.std(abs_diffs_fT[key][1])
        i = 0    
        for cells in ws[f'{columns[key][0]}{l + 3}:{columns[key][1]}{l + 5}']:
            if i == 0:
                cells[0].alignment = Alignment(horizontal='center')
                cells[0].value = mean_abs_diffs_fT[key][0]
                cells[1].alignment = Alignment(horizontal='center')
                cells[1].value = mean_abs_diffs_fT[key][1]
            elif i == 1:
                cells[0].alignment = Alignment(horizontal='center')
                cells[0].value = var_abs_diffs_fT[key][0]
                cells[1].alignment = Alignment(horizontal='center')
                cells[1].value = var_abs_diffs_fT[key][1]
            elif i == 2:
                cells[0].alignment = Alignment(horizontal='center')
                cells[0].value = std_abs_diffs_fT[key][0]
                cells[1].alignment = Alignment(horizontal='center')
                cells[1].value = std_abs_diffs_fT[key][1]
            i += 1
    stat_names = ['Mean of |fT_ - F(mode)|', 'Variance of |fT_ - F(mode)|', 'Standard deviation of |fT_ - F(mode)|']
    b1 = Side(border_style="medium", color="000000")
    k = 0
    ws.column_dimensions['N'].width = np.max([len(stat_name) for stat_name in stat_names])
    for cells in ws[f'N{l + 3}:V{l + 5}']:
        cells[0].alignment = Alignment(horizontal='right')
        cells[0].value = stat_names[k]
        for i in range(1, len(cells)):
            if (i % 2 != 0):
                cells[i].border = Border(left=b1)
        k+=1
     
    